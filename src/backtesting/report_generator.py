"""
Report Generator Module
バックテスト・予測レポート生成モジュール
"""

import pandas as pd
import numpy as np
from typing import Dict, Optional, List
from datetime import datetime
import os
from ..utils.logger import get_logger
from ..utils.config import get_config

logger = get_logger(__name__)

# Excel出力用ライブラリ
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. Excel report generation will be disabled.")

# PDF出力用ライブラリ
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab not available. PDF report generation will be disabled.")


class ReportGenerator:
    """レポート生成クラス"""

    def __init__(self, output_dir: Optional[str] = None):
        """
        初期化

        Args:
            output_dir: 出力ディレクトリ（省略時は設定ファイルから取得）
        """
        self.output_dir = output_dir or get_config('reports.output_dir', 'reports')

        # 出力ディレクトリの作成
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_backtest_report_excel(
        self,
        results: Dict,
        symbol: str,
        filename: Optional[str] = None
    ) -> str:
        """
        バックテストレポートをExcel形式で生成

        Args:
            results: バックテスト結果
            symbol: 銘柄コード
            filename: 出力ファイル名（省略時は自動生成）

        Returns:
            str: 生成されたファイルパス
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl is not installed. Cannot generate Excel report.")
            return ""

        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"backtest_report_{symbol}_{timestamp}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        logger.info(f"Generating backtest report for {symbol}")

        try:
            # Excelワークブックの作成
            wb = openpyxl.Workbook()

            # サマリーシート
            self._create_summary_sheet(wb, results, symbol)

            # ポートフォリオ履歴シート
            if 'portfolio_history' in results:
                self._create_portfolio_history_sheet(wb, results['portfolio_history'])

            # 取引履歴シート
            if 'trades' in results:
                self._create_trades_sheet(wb, results['trades'])

            # 保存
            wb.save(filepath)
            logger.info(f"Excel report saved to {filepath}")

            return filepath

        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return ""

    def _create_summary_sheet(self, wb, results: Dict, symbol: str):
        """サマリーシートの作成"""
        ws = wb.active
        ws.title = "Summary"

        # ヘッダースタイル
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)

        # タイトル
        ws['A1'] = f'Backtest Report - {symbol}'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')

        # 生成日時
        ws['A2'] = 'Generated at:'
        ws['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # パフォーマンスメトリクス
        row = 4
        ws[f'A{row}'] = 'Performance Metrics'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')

        metrics = [
            ('Initial Value', results.get('initial_value', 0), '¥'),
            ('Final Value', results.get('final_value', 0), '¥'),
            ('Total Return', results.get('total_return_pct', 0), '%'),
            ('Sharpe Ratio', results.get('sharpe_ratio', 0), ''),
            ('Max Drawdown', results.get('max_drawdown_pct', 0), '%'),
            ('Win Rate', results.get('win_rate_pct', 0), '%'),
            ('Number of Trades', results.get('num_trades', 0), ''),
        ]

        row += 1
        for metric_name, value, unit in metrics:
            ws[f'A{row}'] = metric_name
            if isinstance(value, float):
                ws[f'B{row}'] = f"{value:,.2f}{unit}"
            else:
                ws[f'B{row}'] = f"{value}{unit}"
            row += 1

        # 列幅の調整
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_portfolio_history_sheet(self, wb, portfolio_df: pd.DataFrame):
        """ポートフォリオ履歴シートの作成"""
        ws = wb.create_sheet(title="Portfolio History")

        # ヘッダー
        headers = ['Date', 'Portfolio Value', 'Cash', 'Positions Value']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)

        # データ
        for row_idx, (_, row) in enumerate(portfolio_df.iterrows(), start=2):
            ws.cell(row=row_idx, column=1, value=str(row['date']))
            ws.cell(row=row_idx, column=2, value=row['value'])
            ws.cell(row=row_idx, column=3, value=row['cash'])
            ws.cell(row=row_idx, column=4, value=row['positions_value'])

        # 列幅の調整
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 18

        # チャートの追加
        if len(portfolio_df) > 0:
            chart = LineChart()
            chart.title = "Portfolio Value Over Time"
            chart.y_axis.title = "Value (¥)"
            chart.x_axis.title = "Time"

            data = Reference(ws, min_col=2, min_row=1, max_row=len(portfolio_df) + 1)
            chart.add_data(data, titles_from_data=True)

            ws.add_chart(chart, "F2")

    def _create_trades_sheet(self, wb, trades_df: pd.DataFrame):
        """取引履歴シートの作成"""
        ws = wb.create_sheet(title="Trade History")

        if trades_df.empty:
            ws['A1'] = 'No trades executed'
            return

        # ヘッダー
        headers = list(trades_df.columns)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)

        # データ
        for row_idx, (_, row) in enumerate(trades_df.iterrows(), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 列幅の調整
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

    def generate_prediction_report_excel(
        self,
        predictions_df: pd.DataFrame,
        symbol: str,
        filename: Optional[str] = None
    ) -> str:
        """
        予測レポートをExcel形式で生成

        Args:
            predictions_df: 予測結果データフレーム
            symbol: 銘柄コード
            filename: 出力ファイル名（省略時は自動生成）

        Returns:
            str: 生成されたファイルパス
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl is not installed. Cannot generate Excel report.")
            return ""

        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"prediction_report_{symbol}_{timestamp}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        logger.info(f"Generating prediction report for {symbol}")

        try:
            # Excelワークブックの作成
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Predictions"

            # タイトル
            ws['A1'] = f'Prediction Report - {symbol}'
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:D1')

            # ヘッダー
            row = 3
            headers = list(predictions_df.columns)
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)

            # データ
            for row_idx, (_, row) in enumerate(predictions_df.iterrows(), start=row + 1):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # 列幅の調整
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 18

            # 保存
            wb.save(filepath)
            logger.info(f"Excel report saved to {filepath}")

            return filepath

        except Exception as e:
            logger.error(f"Error generating prediction Excel report: {e}")
            return ""

    def generate_backtest_report_pdf(
        self,
        results: Dict,
        symbol: str,
        filename: Optional[str] = None
    ) -> str:
        """
        バックテストレポートをPDF形式で生成

        Args:
            results: バックテスト結果
            symbol: 銘柄コード
            filename: 出力ファイル名（省略時は自動生成）

        Returns:
            str: 生成されたファイルパス
        """
        if not REPORTLAB_AVAILABLE:
            logger.error("reportlab is not installed. Cannot generate PDF report.")
            return ""

        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"backtest_report_{symbol}_{timestamp}.pdf"

        filepath = os.path.join(self.output_dir, filename)

        logger.info(f"Generating PDF backtest report for {symbol}")

        try:
            # PDFドキュメントの作成
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()

            # タイトル
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30
            )
            title = Paragraph(f"Backtest Report - {symbol}", title_style)
            story.append(title)

            # 生成日時
            date_text = Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
            story.append(date_text)
            story.append(Spacer(1, 20))

            # パフォーマンスサマリー
            summary_title = Paragraph("Performance Summary", styles['Heading2'])
            story.append(summary_title)
            story.append(Spacer(1, 10))

            # サマリーテーブル
            summary_data = [
                ['Metric', 'Value'],
                ['Initial Value', f"¥{results.get('initial_value', 0):,.0f}"],
                ['Final Value', f"¥{results.get('final_value', 0):,.0f}"],
                ['Total Return', f"{results.get('total_return_pct', 0):.2f}%"],
                ['Sharpe Ratio', f"{results.get('sharpe_ratio', 0):.2f}"],
                ['Max Drawdown', f"{results.get('max_drawdown_pct', 0):.2f}%"],
                ['Win Rate', f"{results.get('win_rate_pct', 0):.2f}%"],
                ['Number of Trades', f"{results.get('num_trades', 0)}"],
            ]

            summary_table = Table(summary_data, colWidths=[8*cm, 8*cm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            story.append(summary_table)

            # PDFのビルド
            doc.build(story)
            logger.info(f"PDF report saved to {filepath}")

            return filepath

        except Exception as e:
            logger.error(f"Error generating PDF report: {e}")
            return ""

    def generate_summary_report(
        self,
        symbol: str,
        backtest_results: Optional[Dict] = None,
        predictions: Optional[pd.DataFrame] = None,
        news_sentiment: Optional[Dict] = None,
        format: str = 'excel'
    ) -> str:
        """
        総合レポートの生成

        Args:
            symbol: 銘柄コード
            backtest_results: バックテスト結果
            predictions: 予測結果
            news_sentiment: ニュース感情分析結果
            format: 出力形式 ('excel' or 'pdf')

        Returns:
            str: 生成されたファイルパス
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"summary_report_{symbol}_{timestamp}.{format}"

        if format == 'excel' and OPENPYXL_AVAILABLE:
            return self._generate_summary_excel(symbol, backtest_results, predictions, news_sentiment, filename)
        elif format == 'pdf' and REPORTLAB_AVAILABLE:
            return self._generate_summary_pdf(symbol, backtest_results, predictions, news_sentiment, filename)
        else:
            logger.error(f"Unsupported format or required library not available: {format}")
            return ""

    def _generate_summary_excel(
        self,
        symbol: str,
        backtest_results: Optional[Dict],
        predictions: Optional[pd.DataFrame],
        news_sentiment: Optional[Dict],
        filename: str
    ) -> str:
        """総合レポート（Excel版）"""
        filepath = os.path.join(self.output_dir, filename)

        try:
            wb = openpyxl.Workbook()

            # バックテスト結果
            if backtest_results:
                self._create_summary_sheet(wb, backtest_results, symbol)

            # 予測結果
            if predictions is not None and not predictions.empty:
                ws = wb.create_sheet(title="Predictions")
                # データの書き込み
                for r_idx, row in enumerate(predictions.itertuples(index=False), start=1):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

            # ニュース感情
            if news_sentiment:
                ws = wb.create_sheet(title="News Sentiment")
                row = 1
                for key, value in news_sentiment.items():
                    ws.cell(row=row, column=1, value=key)
                    ws.cell(row=row, column=2, value=value)
                    row += 1

            wb.save(filepath)
            logger.info(f"Summary Excel report saved to {filepath}")

            return filepath

        except Exception as e:
            logger.error(f"Error generating summary Excel report: {e}")
            return ""

    def _generate_summary_pdf(
        self,
        symbol: str,
        backtest_results: Optional[Dict],
        predictions: Optional[pd.DataFrame],
        news_sentiment: Optional[Dict],
        filename: str
    ) -> str:
        """総合レポート（PDF版）"""
        filepath = os.path.join(self.output_dir, filename)

        try:
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()

            # タイトル
            title = Paragraph(f"Summary Report - {symbol}", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 20))

            # バックテスト結果
            if backtest_results:
                bt_title = Paragraph("Backtest Results", styles['Heading2'])
                story.append(bt_title)
                # テーブル追加...

            # 予測結果
            if predictions is not None:
                pred_title = Paragraph("Predictions", styles['Heading2'])
                story.append(pred_title)
                # テーブル追加...

            # ニュース感情
            if news_sentiment:
                news_title = Paragraph("News Sentiment", styles['Heading2'])
                story.append(news_title)
                # テーブル追加...

            doc.build(story)
            logger.info(f"Summary PDF report saved to {filepath}")

            return filepath

        except Exception as e:
            logger.error(f"Error generating summary PDF report: {e}")
            return ""
